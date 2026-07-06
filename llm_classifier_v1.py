import os
import re
import time
import random

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Το κλειδί διαβάζεται από environment variable, δεν αποθηκεύεται στον κώδικα.
#   Windows (PowerShell):  $env:OPENROUTER_API_KEY="sk-or-..."
#   Windows (cmd):         set OPENROUTER_API_KEY=sk-or-...
API_KEY = os.environ.get("OPENROUTER_API_KEY")
if not API_KEY:
    raise SystemExit(
        "Δεν βρέθηκε το OPENROUTER_API_KEY. Όρισέ το ως environment variable "
        "πριν την εκτέλεση."
    )

MODEL = "openai/gpt-4o-mini"
INPUT_FILE = "sample_300_reporting_obligations.xlsx"
OUTPUT_FILE = "dataset_300_reporting_obligations_output.xlsx"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_color_warning_shown = False

# ======================================================
# PROMPT
# ======================================================
PROMPT = """
Είσαι αυστηρός ταξινομητής νομικών προτάσεων.

Ορισμός:
Ως Υποχρέωση Αναφοράς (Reporting Obligation – RO) ορίζεται η ρητά προβλεπόμενη από το δίκαιο υποχρέωση, βάσει της οποίας ένας καθορισμένος υπόχρεος ή ρυθμιζόμενος φορέας οφείλει να συλλέγει και να παρέχει, σε προκαθορισμένη μορφή, περιεχόμενο και χρονική συχνότητα, συγκεκριμένες πληροφορίες, δεδομένα, δηλώσεις ή εκθέσεις προς αρμόδια δημόσια, ρυθμιστική, εποπτική ή ελεγκτική αρχή ή προς θεσμικά πληροφοριακά συστήματα της διοίκησης, με σκοπό την άσκηση εποπτείας, την παρακολούθηση της συμμόρφωσης ή την εφαρμογή και επιβολή της νομοθεσίας.
Οι υποχρεώσεις αναφοράς καλύπτουν:
- αναφορές από ιδιωτικούς φορείς προς δημόσιες αρχές,
- θεσμικές αναφορές μεταξύ δημόσιων φορέων στο πλαίσιο διοικητικής ή ρυθμιστικής ιεραρχίας,
- καθώς και διασυνοριακές ή υπερεθνικές αναφορές (π.χ. μεταξύ κρατών μελών και ευρωπαϊκών θεσμών).
Βασικά χαρακτηριστικά μιας υποχρέωσης αναφοράς αποτελούν:
- η ύπαρξη νομικά καθορισμένου υπόχρεου,
- η υποχρεωτικότητα της παροχής πληροφορίας,
- ο προσδιορισμός του αποδέκτη (αρμόδια αρχή),
- και ο εποπτικός ή ελεγκτικός σκοπός της αναφοράς.
Η έννοια δεν περιλαμβάνει:
- υποχρεώσεις συμπεριφοράς που δεν συνεπάγονται παροχή πληροφοριών,
- υποχρεώσεις δημοσιοποίησης προς το ευρύ κοινό,
- αιτήσεις ή δηλώσεις που υποβάλλονται κατόπιν πρωτοβουλίας του ενδιαφερομένου για την απόκτηση δικαιώματος ή άδειας,
- ούτε μορφές διοικητικού συντονισμού μεταξύ αρχών που δεν συνδέονται με υποχρέωση εποπτικής αναφοράς.

Παραδείγματα RO 
- Κάθε φορέας της Γενικής Κυβέρνησης υποχρεούται να υποβάλει μηνιαία έκθεση στο αρμόδιο Υπουργείο με τα χρηματοοικονομικά στοιχεία
- Οι εγγεγραμμένοι φορείς υποχρεούνται να αποστέλλουν ετησίως κατάλογο έργων και προϋπολογισμών σε αρμόδια δημόσια υπηρεσία
- Ο αγοραστής φορολογικής ταμειακής μηχανής υποχρεούται να υποβάλλει δήλωση στη Δ.Ο.Υ. με τα στοιχεία απόκτησης και εγκατάστασης

Αντιπαραδείγματα RO
- Οι τελωνειακές αρχές μπορούν να διενεργούν ελέγχους και οι εξαγωγείς οφείλουν να τους επιτρέπουν
- Οι εγγυητικές επιστολές πρέπει να περιλαμβάνουν συγκεκριμένα στοιχεία και όρους για να γίνουν αποδεκτές
- Ο ενδιαφερόμενος μπορεί να αναθέτει καθήκοντα σε άλλο φορέα εφόσον δεν διαθέτει επαρκή τεχνική στελέχωση

---

Δίνεται το παρακάτω απόσπασμα με τρεις διαδοχικές προτάσεις:

[Προηγούμενη πρόταση]: {prev_sentence}
[Υπό εξέταση πρόταση]: {sentence}
[Επόμενη πρόταση]: {next_sentence}

Ταξινόμησε ΜΟΝΟ την [Υπό εξέταση πρόταση]. Οι υπόλοιπες δίνονται μόνο ως πλαίσιο.

Απάντησε μόνο με:
RO - σύντομη αιτιολόγηση
ή
NOT_RO - σύντομη αιτιολόγηση
"""


def parse_llm_output(text: str) -> tuple[str, str]:
    """
    Εξάγει (label, reason) από την απάντηση του μοντέλου. Πιάνει RO/NOT_RO στην αρχή ανεξάρτητα από τον διαχωριστή
    (-, –, —, :, νέα γραμμή) ή τυχόν markdown (**RO**).
    """
    cleaned = text.strip().lstrip("*# ").strip()
    m = re.match(r'^\**\s*(NOT[_\s]?RO|RO)\**\s*[-–—:.\n]*\s*(.*)',
                 cleaned, re.IGNORECASE | re.DOTALL)
    if m:
        label = normalize_label(m.group(1))
        reason = m.group(2).strip().replace("\n", " ")
        return label, reason
    # Fallback: δεν αναγνωρίστηκε καθαρή ετικέτα
    return "UNPARSED", cleaned.replace("\n", " ")


def normalize_label(label: str) -> str:
    """NOT_RO / NOT RO → 'NOT RO', RO → 'RO' (case-insensitive)."""
    norm = label.strip().upper().replace("_", " ")
    norm = re.sub(r'\s+', ' ', norm)
    return norm


def classify(sentence, prev_sentence="", next_sentence=""):
    prompt = PROMPT.format(
        prev_sentence=prev_sentence if prev_sentence else "(δεν υπάρχει)",
        sentence=sentence,
        next_sentence=next_sentence if next_sentence else "(δεν υπάρχει)",
    )

    for attempt in range(5):
        try:
            response = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": MODEL,
                    "temperature": 0,
                    "max_tokens": 100,
                    "messages": [{"role": "user", "content": prompt}],
                },
                timeout=60,
            )

            if response.status_code == 429:
                wait = 5 * (attempt + 1)
                print(f"Rate limited. Waiting {wait}s...")
                time.sleep(wait)
                continue

            data = response.json()
            if "choices" not in data:
                print("API ERROR:", data)
                time.sleep(3)
                continue

            return data["choices"][0]["message"]["content"].strip()

        except Exception as e:
            print("Exception:", e)
            time.sleep(3)

    return "ERROR"


def compute_metrics(df):
    """
    Confusion matrix και precision/recall/F1 για την κλάση RO,
    πάνω σε όλες τις γραμμές που έχουν και πρόβλεψη και ground truth.
    Επιστρέφει dict ή None αν λείπει το ground truth.
    """
    if "Evaluation" not in df.columns:
        return None

    tp = fp = tn = fn = 0
    for _, row in df.iterrows():
        pred_raw = row.get("ChatGPT_Result")
        gold_raw = row.get("Evaluation")
        if pd.isna(pred_raw) or pd.isna(gold_raw):
            continue

        pred = normalize_label(str(pred_raw))
        gold = normalize_label(str(gold_raw))
        if pred not in ("RO", "NOT RO") or gold not in ("RO", "NOT RO"):
            continue

        if pred == "RO" and gold == "RO":
            tp += 1
        elif pred == "RO" and gold == "NOT RO":
            fp += 1
        elif pred == "NOT RO" and gold == "NOT RO":
            tn += 1
        else:
            fn += 1

    evaluated = tp + fp + tn + fn
    if evaluated == 0:
        return None

    precision = tp / (tp + fp) if (tp + fp) else 0.0
    recall = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = (2 * precision * recall / (precision + recall)
          if (precision + recall) else 0.0)
    accuracy = (tp + tn) / evaluated

    return {
        "tp": tp, "fp": fp, "tn": tn, "fn": fn,
        "evaluated": evaluated,
        "precision": precision, "recall": recall,
        "f1": f1, "accuracy": accuracy,
    }


def save_with_colors(df, filepath, result_col, eval_col):
    """
    Αποθηκεύει το df και χρωματίζει τη result_col: πράσινο αν συμφωνεί με
    το ground truth, κόκκινο αν διαφέρει, χωρίς χρώμα αν λείπει.
    """
    global _color_warning_shown

    df.to_excel(filepath, index=False)

    wb = load_workbook(filepath)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    result_idx = headers.get(result_col)
    eval_idx = headers.get(eval_col)

    if result_idx is None or eval_idx is None:
        if not _color_warning_shown:
            print(f"Σημείωση: δεν βρέθηκαν οι στήλες '{result_col}' ή "
                  f"'{eval_col}' — παράλειψη χρωματισμού.")
            _color_warning_shown = True
        wb.save(filepath)
        return

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        result_cell = row[result_idx - 1]
        eval_cell = row[eval_idx - 1]

        if result_cell.value is None:
            continue

        llm_label = normalize_label(str(result_cell.value))
        gold_label = normalize_label(str(eval_cell.value)) if eval_cell.value else ""

        result_cell.fill = GREEN if llm_label == gold_label else RED

    wb.save(filepath)


def main():
    df = pd.read_excel(INPUT_FILE)
    n = len(df)

    if os.path.exists(OUTPUT_FILE):
        df_existing = pd.read_excel(OUTPUT_FILE)
        df["ChatGPT_Result"] = df_existing.get("ChatGPT_Result")
        df["ChatGPT_Reason"] = df_existing.get("ChatGPT_Reason")
        # Μεταφορά τυχόν χειροκίνητου ground truth από το output.
        if "Evaluation" in df_existing.columns:
            df["Evaluation"] = df_existing["Evaluation"]
        already_done = df["ChatGPT_Result"].notna().sum()
        print(f"Βρέθηκε υπάρχον output. {already_done} γραμμές ήδη έτοιμες.")
    else:
        df["ChatGPT_Result"] = None
        df["ChatGPT_Reason"] = None
        print("Νέο output αρχείο. Ξεκινάμε από την αρχή.")

    for i, row in df.iterrows():
        if pd.notna(df.at[i, "ChatGPT_Result"]):
            continue

        text = row["text"]
        prev_text = row.get("previous_sentence", "")
        next_text = row.get("next_sentence", "")

        llm_output = classify(text, prev_sentence=prev_text, next_sentence=next_text)

        if llm_output == "ERROR":
            label, reason = "ERROR", ""
        else:
            label, reason = parse_llm_output(llm_output)

        df.at[i, "ChatGPT_Result"] = label
        df.at[i, "ChatGPT_Reason"] = reason

        print(f"{i + 1}/{n} | {label} - {reason[:80]}")

        save_with_colors(df, OUTPUT_FILE, "ChatGPT_Result", "Evaluation")
        time.sleep(1 + random.random())

    # Τελικά στατιστικά, υπολογισμένα πάντα πάνω σε ΟΛΟ το df.
    ro_count = (df["ChatGPT_Result"].apply(normalize_label) == "RO").sum()
    not_ro_count = (df["ChatGPT_Result"].apply(normalize_label) == "NOT RO").sum()
    errors = df["ChatGPT_Result"].eq("ERROR").sum()
    unparsed = df["ChatGPT_Result"].eq("UNPARSED").sum()
    total = ro_count + not_ro_count
    ro_pct = (ro_count / total * 100) if total else 0

    print(f"\nDone. Saved to {OUTPUT_FILE}")
    print(f"RO: {ro_count} ({ro_pct:.1f}%) | NOT RO: {not_ro_count} | "
          f"ERROR: {errors} | UNPARSED: {unparsed} | Σύνολο ταξινομημένων: {total}")

    metrics = compute_metrics(df)
    if metrics is None:
        print("\nΔεν υπάρχει στήλη 'Evaluation' με ground truth — "
              "συμπλήρωσέ την στο output για να υπολογιστούν precision/recall/F1.")
    else:
        m = metrics
        print("\n================ ΑΞΙΟΛΟΓΗΣΗ (κλάση RO) ================")
        print(f"Δείγματα με ground truth: {m['evaluated']}")
        print(f"                 Gold RO   Gold NOT_RO")
        print(f"  Pred RO        TP={m['tp']:<6}  FP={m['fp']:<6}")
        print(f"  Pred NOT_RO    FN={m['fn']:<6}  TN={m['tn']:<6}")
        print(f"Precision : {m['precision']:.3f}")
        print(f"Recall    : {m['recall']:.3f}")
        print(f"F1        : {m['f1']:.3f}")
        print(f"Accuracy  : {m['accuracy']:.3f}")
        print("======================================================")


if __name__ == "__main__":
    main()
